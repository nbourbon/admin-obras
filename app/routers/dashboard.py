from decimal import Decimal
from typing import List, Optional
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func, extract
from io import BytesIO
from datetime import datetime

from app.database import get_db
from app.schemas.dashboard import (
    DashboardSummary,
    ExpenseEvolution,
    MonthlyExpense,
    UserPaymentStatus,
    ExpensePaymentStatus,
    ParticipantStatus,
    ExpenseByProvider,
    ExpenseByCategory,
    ExpenseByRubro,
)
from app.schemas.contribution import MemberBalanceResponse, ContributionsByParticipant
from app.utils.dependencies import get_current_user, get_project_from_header
from app.models.user import User
from app.models.expense import Expense
from app.models.payment import ParticipantPayment
from app.models.project import Project
from app.models.project_member import ProjectMember
from app.models.contribution import Contribution, ContributionStatus
from app.services.exchange_rate import fetch_blue_dollar_rate_sync
from app.services.expense_splitter import get_user_payment_summary
from app.services.contribution_manager import get_all_member_balances, get_contributions_by_participant

router = APIRouter(prefix="/dashboard", tags=["Dashboard"])


@router.get("/summary", response_model=DashboardSummary)
async def get_dashboard_summary(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    project: Optional[Project] = Depends(get_project_from_header),
):
    """
    Get overall dashboard summary with totals for the current project.
    Optionally filter by date range (expense_date).
    """
    # Build expense query (exclude deleted expenses)
    expense_query = db.query(
        func.sum(Expense.amount_usd).label("total_usd"),
        func.sum(Expense.amount_ars).label("total_ars"),
        func.count(Expense.id).label("count"),
    ).filter(Expense.is_deleted == False)
    if project:
        expense_query = expense_query.filter(Expense.project_id == project.id)

    # Date filters
    if start_date:
        expense_query = expense_query.filter(Expense.expense_date >= datetime.fromisoformat(start_date))
    if end_date:
        expense_query = expense_query.filter(Expense.expense_date <= datetime.fromisoformat(end_date))

    expense_totals = expense_query.first()

    total_expenses_usd = Decimal(str(expense_totals.total_usd or 0))
    total_expenses_ars = Decimal(str(expense_totals.total_ars or 0))
    expenses_count = expense_totals.count or 0

    # Get payment totals (filter by project through expense, exclude deleted)
    # Only count expense payments, not contribution payments
    paid_query = db.query(
        func.sum(ParticipantPayment.amount_due_usd).label("paid_usd"),
        func.sum(ParticipantPayment.amount_due_ars).label("paid_ars"),
    ).join(Expense).filter(
        ParticipantPayment.expense_id.isnot(None),  # Only expense payments
        ParticipantPayment.is_paid == True,
        ParticipantPayment.is_deleted == False,
        Expense.is_deleted == False,
    )

    if project:
        paid_query = paid_query.filter(Expense.project_id == project.id)

    # Date filters on payments
    if start_date:
        paid_query = paid_query.filter(Expense.expense_date >= datetime.fromisoformat(start_date))
    if end_date:
        paid_query = paid_query.filter(Expense.expense_date <= datetime.fromisoformat(end_date))

    paid_payments = paid_query.first()

    total_paid_usd = Decimal(str(paid_payments.paid_usd or 0))
    total_paid_ars = Decimal(str(paid_payments.paid_ars or 0))

    # Calculate pending as direct sum of unpaid dues (avoids rounding drift)
    pending_query = db.query(
        func.sum(ParticipantPayment.amount_due_usd).label("pending_usd"),
        func.sum(ParticipantPayment.amount_due_ars).label("pending_ars"),
    ).join(Expense).filter(
        ParticipantPayment.expense_id.isnot(None),
        ParticipantPayment.is_paid == False,
        ParticipantPayment.is_deleted == False,
        Expense.is_deleted == False,
    )
    if project:
        pending_query = pending_query.filter(Expense.project_id == project.id)
    if start_date:
        pending_query = pending_query.filter(Expense.expense_date >= datetime.fromisoformat(start_date))
    if end_date:
        pending_query = pending_query.filter(Expense.expense_date <= datetime.fromisoformat(end_date))

    pending_result = pending_query.first()
    total_pending_usd = Decimal(str(pending_result.pending_usd or 0))
    total_pending_ars = Decimal(str(pending_result.pending_ars or 0))

    # Get participant count
    if project:
        participants_count = (
            db.query(ProjectMember)
            .join(User)
            .filter(ProjectMember.project_id == project.id)
            .filter(ProjectMember.is_active == True)
            .filter(User.is_active == True)
            .filter(ProjectMember.participation_percentage > 0)
            .count()
        )
    else:
        participants_count = (
            db.query(User)
            .filter(User.is_active == True)
            .filter(User.participation_percentage > 0)
            .count()
        )

    # Get currency_mode and project type from project
    currency_mode = getattr(project, 'currency_mode', None) or "DUAL" if project else "DUAL"
    project_type = getattr(project, 'project_type', None) if project else None

    # Extract square_meters and contribution_mode from type_parameters JSON
    square_meters = None
    contribution_mode = None
    if project and project_type and project_type == "construccion":
        type_params = getattr(project, 'type_parameters', None)
        if type_params and isinstance(type_params, dict):
            square_meters = type_params.get('square_meters')
            if square_meters:
                square_meters = Decimal(str(square_meters))
            contribution_mode = type_params.get('contribution_mode', 'both')  # Default to 'both'

    # Get current exchange rate (skip for single-currency projects)
    if currency_mode == "DUAL":
        try:
            current_rate = fetch_blue_dollar_rate_sync()
        except Exception:
            current_rate = Decimal("0")
    else:
        current_rate = Decimal("0")

    # Get contribution totals (approved only)
    total_contributions_usd = Decimal("0")
    total_contributions_ars = Decimal("0")
    total_balance_usd = Decimal("0")
    total_balance_ars = Decimal("0")

    if project:
        # Get total approved contributions
        # Sum based on currency field (Contribution uses generic amount + currency)
        from app.models.contribution import Currency as ContribCurrency
        from sqlalchemy import case

        contributions_query = db.query(
            func.sum(case(
                (Contribution.currency == ContribCurrency.USD, Contribution.amount),
                else_=0
            )).label("total_usd"),
            func.sum(case(
                (Contribution.currency == ContribCurrency.ARS, Contribution.amount),
                else_=0
            )).label("total_ars"),
        ).filter(
            Contribution.project_id == project.id,
            Contribution.status == ContributionStatus.APPROVED,
        )
        contributions_totals = contributions_query.first()
        total_contributions_usd = Decimal(str(contributions_totals.total_usd or 0))
        total_contributions_ars = Decimal(str(contributions_totals.total_ars or 0))

        # Get total member balances
        members_balances = db.query(ProjectMember).filter(
            ProjectMember.project_id == project.id,
            ProjectMember.is_active == True,
        ).all()

        # Sum up balances (for DUAL mode, calculate USD equivalent in real-time)
        if currency_mode == "DUAL" and current_rate > 0:
            total_balance_ars = sum(Decimal(str(m.balance_ars)) for m in members_balances)
            total_balance_usd = (total_balance_ars / current_rate).quantize(Decimal("0.01"))
        else:
            total_balance_usd = sum(Decimal(str(m.balance_usd)) for m in members_balances)
            total_balance_ars = sum(Decimal(str(m.balance_ars)) for m in members_balances)

    # Calculate cost per square meter for construction projects
    cost_per_square_meter_usd = None
    cost_per_square_meter_ars = None
    if project_type == "construccion" and square_meters and square_meters > 0:
        cost_per_square_meter_usd = (total_expenses_usd / square_meters).quantize(Decimal("0.01"))
        cost_per_square_meter_ars = (total_expenses_ars / square_meters).quantize(Decimal("0.01"))

    return DashboardSummary(
        total_expenses_usd=total_expenses_usd,
        total_expenses_ars=total_expenses_ars,
        total_paid_usd=total_paid_usd,
        total_paid_ars=total_paid_ars,
        total_pending_usd=total_pending_usd,
        total_pending_ars=total_pending_ars,
        expenses_count=expenses_count,
        participants_count=participants_count,
        current_exchange_rate=current_rate,
        currency_mode=currency_mode,
        total_contributions_usd=total_contributions_usd,
        total_contributions_ars=total_contributions_ars,
        total_balance_usd=total_balance_usd,
        total_balance_ars=total_balance_ars,
        project_type=project_type,
        square_meters=square_meters,
        cost_per_square_meter_usd=cost_per_square_meter_usd,
        cost_per_square_meter_ars=cost_per_square_meter_ars,
        contribution_mode=contribution_mode,
    )


@router.get("/evolution", response_model=ExpenseEvolution)
async def get_expense_evolution(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    project: Optional[Project] = Depends(get_project_from_header),
):
    """
    Get monthly expense evolution for the current project.
    Optionally filter by date range (expense_date).
    """
    # Group expenses by year and month (exclude deleted)
    query = db.query(
        extract("year", Expense.expense_date).label("year"),
        extract("month", Expense.expense_date).label("month"),
        func.sum(Expense.amount_usd).label("total_usd"),
        func.sum(Expense.amount_ars).label("total_ars"),
        func.count(Expense.id).label("count"),
    ).filter(Expense.is_deleted == False)

    if project:
        query = query.filter(Expense.project_id == project.id)

    # Date filters
    if start_date:
        query = query.filter(Expense.expense_date >= datetime.fromisoformat(start_date))
    if end_date:
        query = query.filter(Expense.expense_date <= datetime.fromisoformat(end_date))

    monthly_data = (
        query
        .group_by(
            extract("year", Expense.expense_date),
            extract("month", Expense.expense_date),
        )
        .order_by(
            extract("year", Expense.expense_date),
            extract("month", Expense.expense_date),
        )
        .all()
    )

    monthly_expenses = []
    cumulative_usd = Decimal("0")
    cumulative_ars = Decimal("0")

    for row in monthly_data:
        monthly_usd = Decimal(str(row.total_usd or 0))
        monthly_ars = Decimal(str(row.total_ars or 0))
        cumulative_usd += monthly_usd
        cumulative_ars += monthly_ars

        monthly_expenses.append(MonthlyExpense(
            year=int(row.year),
            month=int(row.month),
            total_usd=monthly_usd,
            total_ars=monthly_ars,
            expenses_count=row.count,
        ))

    return ExpenseEvolution(
        monthly_data=monthly_expenses,
        cumulative_usd=cumulative_usd,
        cumulative_ars=cumulative_ars,
    )


@router.get("/my-status", response_model=UserPaymentStatus)
async def get_my_payment_status(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    project: Optional[Project] = Depends(get_project_from_header),
):
    """
    Get current user's payment status summary for the current project.
    """
    project_id = project.id if project else None
    summary = get_user_payment_summary(db, current_user.id, project_id)

    # Get participation percentage and contribution balance from project member
    participation_percentage = Decimal(str(current_user.participation_percentage))
    balance_aportes_usd = Decimal("0")
    balance_aportes_ars = Decimal("0")
    has_pending_contribution = False

    if project:
        member = (
            db.query(ProjectMember)
            .filter(ProjectMember.project_id == project.id)
            .filter(ProjectMember.user_id == current_user.id)
            .first()
        )
        if member:
            participation_percentage = Decimal(str(member.participation_percentage))
            balance_aportes_ars = Decimal(str(member.balance_ars))

            # Calculate USD equivalent based on currency mode
            currency_mode = getattr(project, 'currency_mode', 'DUAL') or 'DUAL'
            if currency_mode == "DUAL":
                # Convert ARS to USD at current rate
                try:
                    current_rate = fetch_blue_dollar_rate_sync()
                    if current_rate > 0:
                        balance_aportes_usd = (balance_aportes_ars / current_rate).quantize(Decimal("0.01"))
                except Exception:
                    balance_aportes_usd = Decimal("0")
            elif currency_mode == "USD":
                balance_aportes_usd = Decimal(str(member.balance_usd))
                balance_aportes_ars = Decimal("0")
            # else ARS mode: balance_aportes_usd stays 0

        # Check for pending contribution payments
        from app.models.contribution_payment import ContributionPayment
        from app.models.contribution import Contribution

        pending_contrib = (
            db.query(ContributionPayment)
            .join(Contribution)
            .filter(
                ContributionPayment.user_id == current_user.id,
                Contribution.project_id == project.id,
                ContributionPayment.is_paid == False,
            )
            .first()
        )
        has_pending_contribution = pending_contrib is not None

    return UserPaymentStatus(
        user_id=current_user.id,
        user_name=current_user.full_name,
        participation_percentage=participation_percentage,
        total_due_usd=summary["total_due_usd"],
        total_due_ars=summary["total_due_ars"],
        total_paid_usd=summary["total_paid_usd"],
        total_paid_ars=summary["total_paid_ars"],
        pending_usd=summary["pending_usd"],
        pending_ars=summary["pending_ars"],
        pending_payments_count=summary["pending_payments_count"],
        balance_aportes_usd=balance_aportes_usd,
        balance_aportes_ars=balance_aportes_ars,
        has_pending_contribution=has_pending_contribution,
    )


@router.get("/all-users-status", response_model=List[UserPaymentStatus])
async def get_all_users_payment_status(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Get payment status for all active participants.
    """
    users = (
        db.query(User)
        .filter(User.is_active == True)
        .filter(User.participation_percentage > 0)
        .all()
    )

    result = []
    for user in users:
        summary = get_user_payment_summary(db, user.id)
        result.append(UserPaymentStatus(
            user_id=user.id,
            user_name=user.full_name,
            participation_percentage=Decimal(str(user.participation_percentage)),
            total_due_usd=summary["total_due_usd"],
            total_due_ars=summary["total_due_ars"],
            total_paid_usd=summary["total_paid_usd"],
            total_paid_ars=summary["total_paid_ars"],
            pending_usd=summary["pending_usd"],
            pending_ars=summary["pending_ars"],
            pending_payments_count=summary["pending_payments_count"],
        ))

    return result


@router.get("/expense-status/{expense_id}", response_model=ExpensePaymentStatus)
async def get_expense_payment_status(
    expense_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Get detailed payment status for a specific expense.
    Shows which participants have paid and which are pending.
    """
    expense = db.query(Expense).filter(Expense.id == expense_id).first()
    if not expense:
        from fastapi import HTTPException, status
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Expense not found",
        )

    payments = (
        db.query(ParticipantPayment)
        .filter(
            ParticipantPayment.expense_id == expense_id,
            ParticipantPayment.is_deleted == False,
        )
        .options(joinedload(ParticipantPayment.user))
        .all()
    )

    participants = []
    paid_count = 0
    pending_count = 0

    for payment in payments:
        user = payment.user
        participants.append(ParticipantStatus(
            payment_id=payment.id,
            user_id=payment.user_id,
            user_name=user.full_name if user else "Unknown",
            amount_due_usd=Decimal(str(payment.amount_due_usd)),
            amount_due_ars=Decimal(str(payment.amount_due_ars)),
            is_pending_approval=payment.is_pending_approval,
            is_paid=payment.is_paid,
            paid_at=payment.paid_at,
            submitted_at=payment.submitted_at,
            rejection_reason=payment.rejection_reason,
            exchange_rate_at_payment=Decimal(str(payment.exchange_rate_at_payment)) if payment.exchange_rate_at_payment else None,
            amount_paid_usd=Decimal(str(payment.amount_paid_usd)) if payment.amount_paid_usd else None,
            amount_paid_ars=Decimal(str(payment.amount_paid_ars)) if payment.amount_paid_ars else None,
            receipt_file_path=payment.receipt_file_path,
        ))

        if payment.is_paid:
            paid_count += 1
        else:
            pending_count += 1

    return ExpensePaymentStatus(
        expense_id=expense.id,
        description=expense.description,
        total_amount_usd=Decimal(str(expense.amount_usd)),
        total_amount_ars=Decimal(str(expense.amount_ars)),
        participants=participants,
        fully_paid=pending_count == 0,
        paid_count=paid_count,
        pending_count=pending_count,
    )


@router.get("/export-excel")
async def export_project_excel(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    project: Optional[Project] = Depends(get_project_from_header),
):
    """
    Export complete project report to Excel with multiple sheets:
    - Sheet 1: All expenses with payment details
    - Sheet 2: Dashboard summary statistics
    - Sheet 3: Summary by participant
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    if not project:
        from fastapi import HTTPException, status
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Project context required (X-Project-ID header)"
        )

    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center_align = Alignment(horizontal="center", vertical="center")
    currency_format = '"$"#,##0.00'
    date_format = "DD/MM/YYYY"
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Determine currency mode
    currency_mode = getattr(project, 'currency_mode', None) or "DUAL"

    # === SHEET 1: GASTOS ===
    ws_expenses = wb.create_sheet("Gastos")

    # Headers adapt to currency mode
    if currency_mode == "ARS":
        headers = ["ID", "Fecha", "Descripción", "Proveedor", "Categoría",
                   "Monto ARS", "Estado", "Pagado", "Pendiente"]
    elif currency_mode == "USD":
        headers = ["ID", "Fecha", "Descripción", "Proveedor", "Categoría",
                   "Monto USD", "Estado", "Pagado", "Pendiente"]
    else:
        headers = ["ID", "Fecha", "Descripción", "Proveedor", "Categoría",
                   "Monto USD", "Monto ARS", "Estado", "Pagado", "Pendiente"]
    ws_expenses.append(headers)

    # Style header row
    for col_num, header in enumerate(headers, 1):
        cell = ws_expenses.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Get expenses (exclude deleted)
    expenses = db.query(Expense).filter(
        Expense.project_id == project.id,
        Expense.is_deleted == False,
    ).order_by(Expense.expense_date.desc()).all()

    # Batch load all payments for the project to avoid N+1
    all_expense_payments = db.query(ParticipantPayment).join(Expense).filter(
        Expense.project_id == project.id,
        ParticipantPayment.is_deleted == False,
    ).all()
    payments_by_expense_id = {}
    for p in all_expense_payments:
        payments_by_expense_id.setdefault(p.expense_id, []).append(p)

    for expense in expenses:
        payments = payments_by_expense_id.get(expense.id, [])

        paid_count = sum(1 for p in payments if p.is_paid)
        pending_count = len(payments) - paid_count
        status = "Pagado" if pending_count == 0 else ("Parcial" if paid_count > 0 else "Pendiente")

        # Remove timezone from date for Excel compatibility
        expense_date = expense.expense_date.replace(tzinfo=None) if expense.expense_date else None

        if currency_mode == "ARS":
            row = [
                expense.id, expense_date, expense.description,
                expense.provider.name if expense.provider else "-",
                expense.category.name if expense.category else "-",
                float(expense.amount_ars), status, paid_count, pending_count,
            ]
        elif currency_mode == "USD":
            row = [
                expense.id, expense_date, expense.description,
                expense.provider.name if expense.provider else "-",
                expense.category.name if expense.category else "-",
                float(expense.amount_usd), status, paid_count, pending_count,
            ]
        else:
            row = [
                expense.id, expense_date, expense.description,
                expense.provider.name if expense.provider else "-",
                expense.category.name if expense.category else "-",
                float(expense.amount_usd), float(expense.amount_ars),
                status, paid_count, pending_count,
            ]
        ws_expenses.append(row)

    # Format columns
    for row in ws_expenses.iter_rows(min_row=2, max_row=ws_expenses.max_row):
        row[0].alignment = center_align  # ID
        row[1].number_format = date_format  # Fecha
        row[5].number_format = currency_format  # Amount
        if currency_mode == "DUAL":
            row[6].number_format = currency_format  # ARS
            row[7].alignment = center_align  # Estado
            row[8].alignment = center_align  # Pagado
            row[9].alignment = center_align  # Pendiente
        else:
            row[6].alignment = center_align  # Estado
            row[7].alignment = center_align  # Pagado
            row[8].alignment = center_align  # Pendiente
        for cell in row:
            cell.border = thin_border

    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        ws_expenses.column_dimensions[get_column_letter(col)].width = 15

    # === SHEET 2: DASHBOARD RESUMEN ===
    ws_summary = wb.create_sheet("Resumen Dashboard")

    # Get summary data
    expense_totals = db.query(
        func.sum(Expense.amount_usd).label("total_usd"),
        func.sum(Expense.amount_ars).label("total_ars"),
        func.count(Expense.id).label("count"),
    ).filter(Expense.project_id == project.id).first()

    from sqlalchemy import Integer

    payment_totals = db.query(
        func.count(ParticipantPayment.id).label("total_payments"),
        func.sum(func.cast(ParticipantPayment.is_paid, Integer)).label("paid_count"),
    ).join(Expense).filter(Expense.project_id == project.id).first()

    total_usd = float(expense_totals.total_usd or 0)
    total_ars = float(expense_totals.total_ars or 0)
    expense_count = expense_totals.count or 0
    total_payments = payment_totals.total_payments or 0
    paid_payments = int(payment_totals.paid_count or 0)
    pending_payments = total_payments - paid_payments

    # Add summary data
    ws_summary.append(["RESUMEN DEL PROYECTO", project.name])
    ws_summary.append(["Fecha de generación", datetime.now().strftime("%d/%m/%Y %H:%M")])
    ws_summary.append([])

    ws_summary.append(["GASTOS"])
    ws_summary.append(["Total de gastos", expense_count])
    ws_summary.append(["Total en USD", total_usd])
    ws_summary.append(["Total en ARS", total_ars])
    ws_summary.append([])

    ws_summary.append(["PAGOS"])
    ws_summary.append(["Total de pagos", total_payments])
    ws_summary.append(["Pagos realizados", paid_payments])
    ws_summary.append(["Pagos pendientes", pending_payments])

    # Style summary sheet
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 20

    # Bold headers
    for row in [1, 4, 9]:
        ws_summary.cell(row=row, column=1).font = Font(bold=True, size=12)

    # Currency format
    ws_summary.cell(row=6, column=2).number_format = currency_format
    ws_summary.cell(row=7, column=2).number_format = currency_format

    # === SHEET 3: POR PARTICIPANTE ===
    ws_participants = wb.create_sheet("Por Participante")

    # Headers adapt to currency mode
    if currency_mode == "ARS":
        part_headers = ["Participante", "Email", "% Participación", "Total a pagar ARS",
                        "Pagado ARS", "Pendiente ARS"]
    elif currency_mode == "USD":
        part_headers = ["Participante", "Email", "% Participación", "Total a pagar USD",
                        "Pagado USD", "Pendiente USD"]
    else:
        part_headers = ["Participante", "Email", "% Participación", "Total a pagar USD",
                        "Pagado USD", "Pendiente USD"]
    ws_participants.append(part_headers)

    # Style header
    for col_num, header in enumerate(part_headers, 1):
        cell = ws_participants.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Get participants
    members = db.query(ProjectMember).join(User).filter(
        ProjectMember.project_id == project.id,
        ProjectMember.is_active == True,
        User.is_active == True,
    ).all()

    for member in members:
        user = member.user
        summary = get_user_payment_summary(db, user.id, project.id)

        if currency_mode == "ARS":
            row = [
                user.full_name, user.email, float(member.participation_percentage),
                float(summary["total_due_ars"]), float(summary["total_paid_ars"]),
                float(summary["pending_ars"]),
            ]
        else:
            row = [
                user.full_name, user.email, float(member.participation_percentage),
                float(summary["total_due_usd"]), float(summary["total_paid_usd"]),
                float(summary["pending_usd"]),
            ]
        ws_participants.append(row)

    # Format columns
    for row in ws_participants.iter_rows(min_row=2, max_row=ws_participants.max_row):
        row[2].number_format = '0.00"%"'  # Percentage
        row[2].alignment = center_align
        row[3].number_format = currency_format  # Total USD
        row[4].number_format = currency_format  # Pagado USD
        row[5].number_format = currency_format  # Pendiente USD
        for cell in row:
            cell.border = thin_border

    # Auto-adjust column widths
    ws_participants.column_dimensions['A'].width = 25
    ws_participants.column_dimensions['B'].width = 30
    ws_participants.column_dimensions['C'].width = 18
    for col in [4, 5, 6]:
        ws_participants.column_dimensions[get_column_letter(col)].width = 18

    # === SHEET 4: APORTES ===
    ws_contributions = wb.create_sheet("Aportes")

    # Get contributions data
    contributions_data = get_contributions_by_participant(db, project.id)

    # Get current exchange rate for DUAL mode
    current_tc = Decimal("0")
    if currency_mode == "DUAL":
        try:
            current_tc = fetch_blue_dollar_rate_sync()
        except:
            current_tc = Decimal("1000")

    # Headers adapt to currency mode
    if currency_mode == "ARS":
        contrib_headers = ["Participante", "Email", "% Participación",
                          "Total Aportado ARS", "Saldo Actual ARS", "Cantidad de Aportes"]
    elif currency_mode == "USD":
        contrib_headers = ["Participante", "Email", "% Participación",
                          "Total Aportado USD", "Saldo Actual USD", "Cantidad de Aportes"]
    else:  # DUAL
        contrib_headers = ["Participante", "Email", "% Participación",
                          "Total Aportado USD", "Total Aportado ARS",
                          "Saldo Actual USD", "Saldo Actual ARS", "Cantidad de Aportes"]

    ws_contributions.append(contrib_headers)

    # Style header
    for col_num, header in enumerate(contrib_headers, 1):
        cell = ws_contributions.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Add contribution data for each participant
    for contrib in contributions_data:
        # Get member balance
        member = db.query(ProjectMember).filter(
            ProjectMember.project_id == project.id,
            ProjectMember.user_id == contrib["user_id"],
        ).first()

        if currency_mode == "ARS":
            row = [
                contrib["user_name"], contrib["user_email"],
                float(contrib["participation_percentage"]),
                float(contrib["total_ars"]),
                float(member.balance_ars) if member else 0,
                contrib["contributions_count"],
            ]
        elif currency_mode == "USD":
            row = [
                contrib["user_name"], contrib["user_email"],
                float(contrib["participation_percentage"]),
                float(contrib["total_usd"]),
                float(member.balance_usd) if member else 0,
                contrib["contributions_count"],
            ]
        else:  # DUAL
            # Calculate balance USD from balance ARS
            balance_usd = (member.balance_ars / current_tc).quantize(Decimal("0.01")) if member and current_tc > 0 else Decimal("0")
            row = [
                contrib["user_name"], contrib["user_email"],
                float(contrib["participation_percentage"]),
                float(contrib["total_usd"]),
                float(contrib["total_ars"]),
                float(balance_usd),
                float(member.balance_ars) if member else 0,
                contrib["contributions_count"],
            ]
        ws_contributions.append(row)

    # Format columns
    for row in ws_contributions.iter_rows(min_row=2, max_row=ws_contributions.max_row):
        row[2].number_format = '0.00"%"'  # Percentage
        row[2].alignment = center_align

        if currency_mode == "DUAL":
            row[3].number_format = currency_format  # Total USD
            row[4].number_format = currency_format  # Total ARS
            row[5].number_format = currency_format  # Balance USD
            row[6].number_format = currency_format  # Balance ARS
            row[7].alignment = center_align  # Count
        else:
            row[3].number_format = currency_format  # Total
            row[4].number_format = currency_format  # Balance
            row[5].alignment = center_align  # Count

        for cell in row:
            cell.border = thin_border

    # Auto-adjust column widths
    ws_contributions.column_dimensions['A'].width = 25
    ws_contributions.column_dimensions['B'].width = 30
    ws_contributions.column_dimensions['C'].width = 18
    for col in range(4, len(contrib_headers) + 1):
        ws_contributions.column_dimensions[get_column_letter(col)].width = 18

    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # Generate filename
    filename = f"Reporte_{project.name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"

    # Return as streaming response
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/expenses-by-provider", response_model=List[ExpenseByProvider])
async def get_expenses_by_provider(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    project: Optional[Project] = Depends(get_project_from_header),
):
    """
    Get total expenses grouped by provider, ordered by total descending.
    Optionally filter by date range (expense_date).
    """
    from app.models.provider import Provider
    from sqlalchemy import case

    # Build expense query
    query = db.query(
        Expense.provider_id,
        case(
            (Expense.provider_id.is_(None), "Sin proveedor"),
            else_=Provider.name
        ).label("provider_name"),
        func.sum(Expense.amount_usd).label("total_usd"),
        func.sum(Expense.amount_ars).label("total_ars"),
        func.count(Expense.id).label("expenses_count"),
    ).outerjoin(Provider, Expense.provider_id == Provider.id)

    # Filter by project
    if project:
        query = query.filter(Expense.project_id == project.id)

    # Filter deleted expenses
    query = query.filter(Expense.is_deleted == False)

    # Date filters
    if start_date:
        query = query.filter(Expense.expense_date >= datetime.fromisoformat(start_date))
    if end_date:
        query = query.filter(Expense.expense_date <= datetime.fromisoformat(end_date))

    # Group and order
    query = query.group_by(Expense.provider_id, Provider.name).order_by(func.sum(Expense.amount_usd).desc())

    results = query.all()

    return [
        ExpenseByProvider(
            provider_id=r.provider_id,
            provider_name=r.provider_name,
            total_usd=Decimal(str(r.total_usd or 0)),
            total_ars=Decimal(str(r.total_ars or 0)),
            expenses_count=r.expenses_count,
        )
        for r in results
    ]


@router.get("/expenses-by-category", response_model=List[ExpenseByCategory])
async def get_expenses_by_category(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    project: Optional[Project] = Depends(get_project_from_header),
):
    """
    Get total expenses grouped by category, ordered by total descending.
    Optionally filter by date range (expense_date).
    """
    from app.models.category import Category
    from sqlalchemy import case

    # Build expense query
    query = db.query(
        Expense.category_id,
        case(
            (Expense.category_id.is_(None), "Sin categoría"),
            else_=Category.name
        ).label("category_name"),
        func.sum(Expense.amount_usd).label("total_usd"),
        func.sum(Expense.amount_ars).label("total_ars"),
        func.count(Expense.id).label("expenses_count"),
    ).outerjoin(Category, Expense.category_id == Category.id)

    # Filter by project
    if project:
        query = query.filter(Expense.project_id == project.id)

    # Filter deleted expenses
    query = query.filter(Expense.is_deleted == False)

    # Date filters
    if start_date:
        query = query.filter(Expense.expense_date >= datetime.fromisoformat(start_date))
    if end_date:
        query = query.filter(Expense.expense_date <= datetime.fromisoformat(end_date))

    # Group and order
    query = query.group_by(Expense.category_id, Category.name).order_by(func.sum(Expense.amount_usd).desc())

    results = query.all()

    return [
        ExpenseByCategory(
            category_id=r.category_id,
            category_name=r.category_name,
            total_usd=Decimal(str(r.total_usd or 0)),
            total_ars=Decimal(str(r.total_ars or 0)),
            expenses_count=r.expenses_count,
        )
        for r in results
    ]


@router.get("/expenses-by-rubro", response_model=List[ExpenseByRubro])
async def get_expenses_by_rubro(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    project: Optional[Project] = Depends(get_project_from_header),
):
    """
    Get total expenses grouped by rubro, ordered by total descending.
    Optionally filter by date range (expense_date).
    """
    from app.models.rubro import Rubro
    from sqlalchemy import case

    query = db.query(
        Expense.rubro_id,
        case(
            (Expense.rubro_id.is_(None), "Sin rubro"),
            else_=Rubro.name
        ).label("rubro_name"),
        func.sum(Expense.amount_usd).label("total_usd"),
        func.sum(Expense.amount_ars).label("total_ars"),
        func.count(Expense.id).label("expenses_count"),
    ).outerjoin(Rubro, Expense.rubro_id == Rubro.id)

    if project:
        query = query.filter(Expense.project_id == project.id)

    query = query.filter(Expense.is_deleted == False)

    if start_date:
        query = query.filter(Expense.expense_date >= datetime.fromisoformat(start_date))
    if end_date:
        query = query.filter(Expense.expense_date <= datetime.fromisoformat(end_date))

    query = query.group_by(Expense.rubro_id, Rubro.name).order_by(func.sum(Expense.amount_usd).desc())

    results = query.all()

    return [
        ExpenseByRubro(
            rubro_id=r.rubro_id,
            rubro_name=r.rubro_name,
            total_usd=Decimal(str(r.total_usd or 0)),
            total_ars=Decimal(str(r.total_ars or 0)),
            expenses_count=r.expenses_count,
        )
        for r in results
    ]


@router.get("/balances", response_model=List[MemberBalanceResponse])
async def get_member_balances(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    project: Optional[Project] = Depends(get_project_from_header),
):
    """
    Get balances for all members of the current project.
    For DUAL mode, USD equivalent is calculated in real-time using current exchange rate.
    """
    if not project:
        from fastapi import HTTPException
        raise HTTPException(
            status_code=400,
            detail="X-Project-ID header is required",
        )

    currency_mode = getattr(project, 'currency_mode', 'DUAL') or 'DUAL'
    balances_data = get_all_member_balances(db, project.id)

    return [
        MemberBalanceResponse(
            user_id=b["user_id"],
            user_name=b["user_name"],
            user_email=b["user_email"],
            participation_percentage=b["participation_percentage"],
            balance_usd=b["balance_usd"],
            balance_ars=b["balance_ars"],
            balance_updated_at=b["balance_updated_at"],
        )
        for b in balances_data
    ]


@router.get("/contributions-by-participant", response_model=List[ContributionsByParticipant])
async def get_contributions_by_participant_endpoint(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    project: Optional[Project] = Depends(get_project_from_header),
):
    """
    Get total approved contributions by participant for the current project.
    Shows accumulated historical contributions (for pie chart).
    """
    if not project:
        from fastapi import HTTPException
        raise HTTPException(
            status_code=400,
            detail="X-Project-ID header is required",
        )

    contributions_data = get_contributions_by_participant(db, project.id)

    return [
        ContributionsByParticipant(
            user_id=c["user_id"],
            user_name=c["user_name"],
            user_email=c["user_email"],
            participation_percentage=c["participation_percentage"],
            total_usd=c["total_usd"],
            total_ars=c["total_ars"],
            contributions_count=c["contributions_count"],
        )
        for c in contributions_data
    ]
